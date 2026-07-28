"""Excel audit workbook for predictions and provenance."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.base import FitResult

# Canonical answer for "which values mimic con sordina on IOWA/ORCHIDEA?"
PRIMARY_SHEET = "Predictions_supported"
PRIMARY_COLUMN = "y_pred"


def _highlight_use_this(ws_readme, *, file_name: str) -> None:
    """Bold + yellow highlight for the three fields users must read first."""
    yellow = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True, size=12)
    for row in ws_readme.iter_rows(min_row=1, max_row=ws_readme.max_row, max_col=2):
        key = str(row[0].value or "")
        if key in {
            "1_FILE_NAME",
            "2_PAGE_NAME",
            "3_COLUMN_NAME",
            "USE_THIS_FILE",
            "USE_THIS_SHEET",
            "USE_THIS_COLUMN",
        }:
            for cell in row:
                cell.fill = yellow
                cell.font = bold
    ws_readme["A1"].alignment = Alignment(wrap_text=True)
    ws_readme.column_dimensions["A"].width = 22
    ws_readme.column_dimensions["B"].width = 72


def _highlight_primary_column(ws, column_name: str = PRIMARY_COLUMN) -> None:
    """Highlight the primary prediction column header on Predictions_supported."""
    yellow = PatternFill("solid", fgColor="FFFF00")
    bold = Font(bold=True, size=12)
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for idx, cell in enumerate(header_row, start=1):
        if str(cell.value) == column_name:
            cell.fill = yellow
            cell.font = bold
            # Tint data cells lightly so the column is obvious when scrolling
            light = PatternFill("solid", fgColor="FFF2CC")
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).fill = light
            ws.column_dimensions[get_column_letter(idx)].width = 14
            break


def export_audit_workbook(
    path: str | Path,
    *,
    fit: FitResult,
    bridge: pd.DataFrame,
    target: pd.DataFrame,
    predictions: pd.DataFrame,
    predictions_supported: pd.DataFrame | None = None,
    factor_summary: pd.DataFrame | None = None,
    quality_report: pd.DataFrame | None = None,
    preflight: pd.DataFrame | None = None,
    cv_table: pd.DataFrame | None = None,
    config: dict | None = None,
    audit: dict | None = None,
    model_comparison: pd.DataFrame | None = None,
    calibration: pd.DataFrame | None = None,
    holdout_summary: pd.DataFrame | None = None,
    holdout_detail: pd.DataFrame | None = None,
    sensitivity: pd.DataFrame | None = None,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    file_name = path.name

    n_sup = len(predictions_supported) if predictions_supported is not None else 0
    readme = pd.DataFrame(
        {
            "Field": [
                "1_FILE_NAME",
                "2_PAGE_NAME",
                "3_COLUMN_NAME",
                "how_to_use",
                "estimate_class",
                "scientific_label",
                "model_id",
                "backend",
                "metric",
                "bridge_n",
                "n_predictions_total",
                "n_predictions_supported",
                "primary_sheet",
                "warning",
            ],
            "Value": [
                file_name,
                PRIMARY_SHEET,
                PRIMARY_COLUMN,
                (
                    f"To mimic con sordina (or other special technique) on IOWA/ORCHIDEA: "
                    f"open file [{file_name}] → sheet [{PRIMARY_SHEET}] → column [{PRIMARY_COLUMN}]. "
                    f"Optional uncertainty: y_pred_lo95 / y_pred_hi95."
                ),
                "model_derived_synthetic",
                fit.label,
                fit.model_id,
                fit.backend,
                fit.metric,
                fit.bridge_n,
                len(predictions),
                n_sup,
                PRIMARY_SHEET,
                "Use Predictions_supported!y_pred for robust work. "
                "Predictions_all includes extrapolated dynamics/registers for diagnostics only. "
                "Never label rows as measured IOWA/ORCHIDEA observations. "
                "See METHODOLOGY.md in the tool folder.",
            ],
        }
    )

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        readme.to_excel(xl, sheet_name="README", index=False)
        if config is not None:
            pd.DataFrame([{"key": k, "value": str(v)} for k, v in config.items()]).to_excel(
                xl, sheet_name="Config", index=False
            )
        if preflight is not None:
            preflight.to_excel(xl, sheet_name="Preflight", index=False)
        if quality_report is not None:
            quality_report.to_excel(xl, sheet_name="Quality_report", index=False)
        if cv_table is not None:
            cv_table.to_excel(xl, sheet_name="Blocked_CV", index=False)
        if model_comparison is not None and len(model_comparison):
            model_comparison.to_excel(xl, sheet_name="Model_comparison", index=False)
        if calibration is not None and len(calibration):
            calibration.to_excel(xl, sheet_name="Calibration", index=False)
        if holdout_summary is not None and len(holdout_summary):
            holdout_summary.to_excel(xl, sheet_name="Holdout_summary", index=False)
        if holdout_detail is not None and len(holdout_detail):
            holdout_detail.to_excel(xl, sheet_name="Holdout_detail", index=False)
        if sensitivity is not None and len(sensitivity):
            sensitivity.to_excel(xl, sheet_name="Sensitivity", index=False)
        if audit:
            pd.DataFrame([audit]).to_excel(xl, sheet_name="Audit", index=False)
        bridge.to_excel(xl, sheet_name="Bridge_log_ratios", index=False)
        if factor_summary is not None:
            factor_summary.to_excel(xl, sheet_name="Factor_summary", index=False)
        if fit.effects is not None:
            fit.effects.to_excel(xl, sheet_name="Model_effects", index=False)
        target.to_excel(xl, sheet_name="Target_ordinario", index=False)
        if predictions_supported is not None:
            predictions_supported.to_excel(xl, sheet_name=PRIMARY_SHEET, index=False)
        predictions.to_excel(xl, sheet_name="Predictions_all", index=False)
        needed = {"dynamic", "bridge_dynamic_used", "dynamic_match", "support_level"}
        if needed.issubset(predictions.columns):
            cols = ["dynamic", "bridge_dynamic_used", "dynamic_match", "support_level"]
            if "dynamic_adequate" in predictions.columns:
                cols.append("dynamic_adequate")
            (
                predictions.groupby(cols, dropna=False)
                .size()
                .reset_index(name="n")
                .to_excel(xl, sheet_name="Dynamic_matching", index=False)
            )
        pd.DataFrame(
            [
                {
                    "USE_THIS": f"{file_name} / {PRIMARY_SHEET} / {PRIMARY_COLUMN}",
                    "formula_prediction": "y_pred = y_ordinario * exp(log_effect)",
                    "formula_interval": "y_pred * exp(+/- calibrated_halfwidth_log)",
                    "dynamic_policy": "Supported only for adequate pairs: pp<->{pp,p}, mf<->{mf,mp}, ff<->{ff,f}.",
                    "acoustic_policy": (
                        "Soft literature-aligned priors (not activated EWSD laws) + winsor + shrink + clip. "
                        "Mute is frequency-dependent; scalar log-ratio is an approximation."
                    ),
                    "register_policy": "Outside bridge MIDI: shrunk global effect, not spline extrapolation.",
                    "validation": (
                        "Blocked_CV + Model_comparison + Calibration + Holdout_* + Sensitivity sheets."
                    ),
                }
            ]
        ).to_excel(xl, sheet_name="Formulas", index=False)

        # Visual highlights (after all sheets written)
        _highlight_use_this(xl.book["README"], file_name=file_name)
        if PRIMARY_SHEET in xl.book.sheetnames:
            _highlight_primary_column(xl.book[PRIMARY_SHEET], PRIMARY_COLUMN)
        if "Quality_report" in xl.book.sheetnames:
            yellow = PatternFill("solid", fgColor="FFFF00")
            bold = Font(bold=True, size=12)
            for row in xl.book["Quality_report"].iter_rows(min_row=2, max_col=2):
                if str(row[0].value or "") in {
                    "1_FILE_NAME",
                    "2_PAGE_NAME",
                    "3_COLUMN_NAME",
                }:
                    for cell in row:
                        cell.fill = yellow
                        cell.font = bold
        if "Formulas" in xl.book.sheetnames:
            cell = xl.book["Formulas"]["A2"]
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.font = Font(bold=True, size=12)

    return path
